import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from database import obtener_asistencias_rango, obtener_asistencias_por_fecha

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORTS_FOLDER = os.path.join(BASE_DIR, "reportes")
os.makedirs(REPORTS_FOLDER, exist_ok=True)


def generar_excel_asistencias(fecha_inicio: str, fecha_fin: str = None):
    """
    Genera un archivo Excel con el reporte de asistencias
    
    Args:
        fecha_inicio: Fecha inicial en formato YYYY-MM-DD
        fecha_fin: Fecha final en formato YYYY-MM-DD (opcional)
    
    Returns:
        str: Ruta del archivo generado
    """
    
    # Si no hay fecha_fin, usar solo fecha_inicio (reporte de un día)
    if not fecha_fin:
        fecha_fin = fecha_inicio
        asistencias = obtener_asistencias_por_fecha(fecha_inicio)
        titulo = f"Reporte_Asistencia_{fecha_inicio}"
    else:
        asistencias = obtener_asistencias_rango(fecha_inicio, fecha_fin)
        titulo = f"Reporte_Asistencia_{fecha_inicio}_a_{fecha_fin}"
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencias"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título del reporte
    ws.merge_cells('A1:E1')
    titulo_cell = ws['A1']
    titulo_cell.value = f"REPORTE DE ASISTENCIAS"
    titulo_cell.font = Font(bold=True, size=16, color="4472C4")
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Subtítulo con rango de fechas
    ws.merge_cells('A2:E2')
    subtitulo_cell = ws['A2']
    if fecha_inicio == fecha_fin:
        subtitulo_cell.value = f"Fecha: {fecha_inicio}"
    else:
        subtitulo_cell.value = f"Período: {fecha_inicio} al {fecha_fin}"
    subtitulo_cell.font = Font(size=11)
    subtitulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Fecha de generación
    ws.merge_cells('A3:E3')
    fecha_gen_cell = ws['A3']
    fecha_gen_cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    fecha_gen_cell.font = Font(size=9, italic=True)
    fecha_gen_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados de la tabla
    headers = ["#", "Nombre", "Código", "Fecha", "Hora"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Datos
    if not asistencias:
        ws.merge_cells('A6:E6')
        no_data_cell = ws['A6']
        no_data_cell.value = "No se encontraron registros de asistencia"
        no_data_cell.alignment = Alignment(horizontal="center", vertical="center")
        no_data_cell.font = Font(italic=True, color="999999")
    else:
        for idx, asistencia in enumerate(asistencias, start=1):
            row = idx + 5
            
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=asistencia.get('nombre', 'N/A'))
            ws.cell(row=row, column=3, value=asistencia.get('codigo', 'N/A'))
            ws.cell(row=row, column=4, value=asistencia.get('fecha', 'N/A'))
            ws.cell(row=row, column=5, value=asistencia.get('hora', 'N/A'))
            
            # Aplicar bordes
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border_style
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        
        # Resumen al final
        total_row = len(asistencias) + 7
        ws.merge_cells(f'A{total_row}:D{total_row}')
        total_cell = ws[f'A{total_row}']
        total_cell.value = "TOTAL DE ASISTENCIAS:"
        total_cell.font = Font(bold=True, size=11)
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        total_value_cell = ws[f'E{total_row}']
        total_value_cell.value = len(asistencias)
        total_value_cell.font = Font(bold=True, size=11)
        total_value_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        total_value_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_value_cell.border = border_style
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    
    # Guardar archivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{titulo}_{timestamp}.xlsx"
    filepath = os.path.join(REPORTS_FOLDER, filename)
    
    wb.save(filepath)
    print(f"✅ Reporte generado: {filepath}")
    
    return filepath


def generar_excel_por_estudiante(nombre_estudiante: str):
    """
    Genera un reporte Excel del historial de un estudiante específico
    """
    from database import obtener_historial_estudiante
    
    asistencias = obtener_historial_estudiante(nombre_estudiante)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"
    
    # Título
    ws.merge_cells('A1:C1')
    ws['A1'].value = f"HISTORIAL DE ASISTENCIAS - {nombre_estudiante}"
    ws['A1'].font = Font(bold=True, size=14, color="4472C4")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = ["Fecha", "Hora", "Estado"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Datos
    for idx, asistencia in enumerate(asistencias, start=4):
        ws.cell(row=idx, column=1, value=asistencia.get('fecha', 'N/A'))
        ws.cell(row=idx, column=2, value=asistencia.get('hora', 'N/A'))
        ws.cell(row=idx, column=3, value=asistencia.get('estado', 'N/A'))
    
    # Ajustar columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    
    # Guardar
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"Historial_{nombre_estudiante}_{timestamp}.xlsx"
    filepath = os.path.join(REPORTS_FOLDER, filename)
    
    wb.save(filepath)
    return filepath